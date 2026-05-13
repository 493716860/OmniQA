"""
backend/api/services/importer.py

文件用途
-------
Excel 用例导入服务（ExcelImportService）。

该模块负责把“外部可维护”的 Excel 测试用例，转换成平台内部的数据模型：
Project / Environment / Module / ApiDefinition / ApiCase。

导入规则（可作为面试时讲解的“数据规范”）
--------------------------------------
1) 工作簿要求：
   - 必须包含 data 页签（业务用例）
   - 可选 setup 页签（前置/准备用例，例如登录、初始化数据）
2) 表头要求：
   REQUIRED_HEADERS = {id,title,subtitle,level,url,method,header,payload,expect,rely}
   - 表头大小写不敏感，会统一 lower
3) 字段解析：
   - header / payload / expect 支持 JSON 或 YAML 文本；safe_parse_cell() 负责容错解析
   - url 会 normalize_path()：若是完整 URL 仅取 path，避免把域名写死进用例
4) 写入策略：
   - ApiDefinition：按 (module, path, method) update_or_create
   - ApiCase：按 (api, case_code, is_setup) update_or_create
   - rely：导入后统一回填 dependencies（避免前向引用时找不到依赖）
5) 默认变量提取：
   - default_extractors() 对“登录”类用例提供 token/uid 等默认提取示例

与系统其它模块的关系
-------------------
- Web 层入口：backend/api/views.py -> ExcelImportView 调用此服务
- 执行引擎：backend/api/services/db_runner.py 使用 ApiCase/ApiDefinition 等数据执行
"""

import json
from dataclasses import dataclass, field
from pathlib import PurePosixPath
from urllib.parse import urlparse

import yaml
from openpyxl import load_workbook

from api.models import ApiCase, ApiDefinition, Environment, Module, Project

REQUIRED_HEADERS = {"id", "title", "subtitle", "level", "url", "method", "header", "payload", "expect", "rely"}


@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    failed: int = 0
    errors: list = field(default_factory=list)

    def as_dict(self):
        return {
            "created": self.created,
            "updated": self.updated,
            "failed": self.failed,
            "errors": self.errors,
        }


def safe_parse_cell(value, default):
    if value in (None, ""):
        return default
    if isinstance(value, (dict, list, int, float, bool)):
        return value
    raw = str(value).strip()
    if raw in ("", "None", "null"):
        return default
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        parsed = yaml.safe_load(raw)
        return default if parsed is None else parsed


def normalize_path(url):
    if not url:
        return "/"
    parsed = urlparse(str(url))
    if parsed.scheme and parsed.netloc:
        path = parsed.path or "/"
        return str(PurePosixPath(path))
    return str(url)


def get_cell_map(sheet):
    headers = {}
    for cell in sheet[1]:
        if cell.value is not None:
            key = str(cell.value).strip().lower()
            if key == "subtitle":
                key = "subtitle"
            headers[key] = cell.column
    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValueError(f"缺少表头: {', '.join(sorted(missing))}")
    return headers


def row_value(sheet, headers, row_idx, key):
    value = sheet.cell(row=row_idx, column=headers[key]).value
    return "" if value is None else value


class ExcelImportService:
    def import_file(self, *, file_obj, project_name, module_name="", environment_name="测试环境(dev)", base_url=""):
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        try:
            if "data" not in wb.sheetnames:
                raise ValueError("Excel 必须包含 data 页签")

            project, _ = Project.objects.get_or_create(name=project_name)
            env, _ = Environment.objects.get_or_create(
                project=project,
                name=environment_name,
                defaults={"base_url": base_url},
            )
            if base_url and env.base_url != base_url:
                env.base_url = base_url
                env.save(update_fields=["base_url", "updated_at"])

            stats = ImportStats()
            for sheet_name, is_setup in (("setup", True), ("data", False)):
                if sheet_name not in wb.sheetnames:
                    continue
                self._import_sheet(wb[sheet_name], project, module_name or project_name, stats, is_setup)
            return {"project": project.id, "environment": env.id, **stats.as_dict()}
        finally:
            wb.close()

    def _import_sheet(self, sheet, project, module_name, stats, is_setup):
        headers = get_cell_map(sheet)
        module, _ = Module.objects.get_or_create(project=project, name=module_name)
        pending_dependencies = []
        for row_idx in range(2, sheet.max_row + 1):
            case_code = row_value(sheet, headers, row_idx, "id")
            if not case_code:
                break
            try:
                title = str(row_value(sheet, headers, row_idx, "title") or "")
                method = str(row_value(sheet, headers, row_idx, "method") or "POST").upper()
                raw_url = row_value(sheet, headers, row_idx, "url")
                path = normalize_path(raw_url)
                header = safe_parse_cell(row_value(sheet, headers, row_idx, "header"), {})
                payload = safe_parse_cell(row_value(sheet, headers, row_idx, "payload"), {})
                expect = safe_parse_cell(row_value(sheet, headers, row_idx, "expect"), {})
                if header and not isinstance(header, dict):
                    raise ValueError("Header 必须是 JSON/YAML 对象")

                api, _ = ApiDefinition.objects.update_or_create(
                    module=module,
                    path=path,
                    method=method,
                    defaults={"name": title or path},
                )
                case, created = ApiCase.objects.update_or_create(
                    api=api,
                    case_code=str(case_code),
                    is_setup=is_setup,
                    defaults={
                        "title": title,
                        "subtitle": str(row_value(sheet, headers, row_idx, "subtitle") or ""),
                        "header": header or {},
                        "payload": payload,
                        "expect": expect,
                        "extractors": default_extractors(case_code, title, is_setup),
                        "level": int(row_value(sheet, headers, row_idx, "level") or 1),
                        "rely": str(row_value(sheet, headers, row_idx, "rely") or ""),
                        "sort_order": row_idx,
                        "enabled": True,
                    },
                )
                rely = str(row_value(sheet, headers, row_idx, "rely") or "").strip()
                if rely:
                    pending_dependencies.append((case, parse_rely_codes(rely)))
                stats.created += 1 if created else 0
                stats.updated += 0 if created else 1
            except Exception as exc:
                stats.failed += 1
                stats.errors.append({"sheet": sheet.title, "row": row_idx, "error": str(exc)})

        for case, codes in pending_dependencies:
            dependencies = ApiCase.objects.filter(
                api__module__project=project,
                case_code__in=codes,
            )
            case.dependencies.set(dependencies)


def parse_rely_codes(raw):
    return [item.strip() for item in raw.replace("，", ",").split(",") if item.strip()]


def default_extractors(case_code, title, is_setup):
    if not is_setup and "登录" not in str(title):
        return []
    prefix = str(case_code)
    return [
        {"name": "token", "path": "$.data.token", "required": False},
        {"name": "uid", "path": "$.data.uid", "required": False},
        {"name": "uid_str", "path": "$.data.uid_str", "required": False},
        {"name": "cellphone", "path": "$.data.cellphone", "required": False},
        {"name": f"{prefix}_data", "path": "$.data", "required": False},
    ]
